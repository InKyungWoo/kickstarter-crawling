"""
Phase C: 수집된 JSONL → Excel(카테고리별 시트) 또는 CSV 변환.
"""
import json
import logging
import re
from datetime import datetime, timezone

import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)

# Excel이 허용하지 않는 제어 문자
ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

COLUMNS = [
    "project_url", "title", "blurb", "story_text", "story_truncated", "state",
    "backers_count", "goal", "pledged", "currency",
    "usd_rate_campaign", "usd_pledged", "usd_rate_current", "usd_pledged_current", "percent_funded",
    "has_photos", "photo_count", "has_video", "story_video_count", "has_main_video",
    "start_date", "end_date", "duration_days",
    "creator_name", "creator_badges", "creator_created_count", "creator_backed_count",
    "is_project_we_love", "updates_count", "comments_count",
]


def parse_story(story_html: str | None) -> tuple[str, int, int]:
    """story HTML → (plain text, 이미지 수, 동영상 수)"""
    if not story_html:
        return "", 0, 0
    soup = BeautifulSoup(story_html, "lxml")
    img_count = len(soup.find_all("img"))
    video_count = len(soup.find_all("video")) + len(soup.find_all("iframe"))
    text = soup.get_text("\n", strip=True)
    return text, img_count, video_count


def fmt_date(ts) -> str | None:
    if not ts:
        return None
    return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d")


def build_row(p: dict) -> dict:
    g = p.get("_graph") or {}
    creator_g = g.get("creator") or {}
    story_text, img_count, video_count = parse_story(g.get("story"))

    risks = (g.get("risks") or "").strip()
    if risks:
        story_text = f"{story_text}\n\nRisks and challenges\n{risks}".strip()
    env_texts = [(e.get("description") or "").strip()
                 for e in (g.get("environmentalCommitments") or [])]
    env_texts = [t for t in env_texts if t]
    if env_texts:
        story_text = (story_text + "\n\nEnvironmental commitments\n" + "\n\n".join(env_texts)).strip()
    ai = g.get("aiDisclosure") or {}
    if ai.get("involvesAi"):
        ai_texts = [str(ai[k]).strip() for k in
                    ("generatedByAiDetails", "generatedByAiConsent", "otherAiDetails",
                     "fundingForAiOption", "fundingForAiConsent", "fundingForAiAttribution")
                    if ai.get(k)]
        story_text = (story_text + "\n\nUse of AI\n" + "\n\n".join(ai_texts)).strip()

    badges = list(creator_g.get("badges") or [])
    if creator_g.get("isSuperbacker") and "Superbacker" not in badges:
        badges.append("Superbacker")

    # 후원 목록이 비공개면 backedProjects가 null → backingsCount로 폴백
    backed = (creator_g.get("backedProjects") or {}).get("totalCount")
    if backed is None:
        backed = creator_g.get("backingsCount")

    launched, deadline = p.get("launched_at"), p.get("deadline")
    truncated = len(story_text) > config.EXCEL_CELL_LIMIT

    return {
        "project_url": (p.get("urls") or {}).get("web", {}).get("project"),
        "title": p.get("name"),
        "blurb": p.get("blurb"),
        "story_text": ILLEGAL_CHARS.sub("", story_text)[:config.EXCEL_CELL_LIMIT],
        "story_truncated": truncated,
        "state": p.get("state"),
        "backers_count": g.get("backersCount", p.get("backers_count")),
        "goal": p.get("goal"),
        "pledged": p.get("pledged"),    ## 원금
        "currency": p.get("currency"),  ## 원금 통화
        "usd_rate_campaign": round(float(p["static_usd_rate"]), 4) if p.get("static_usd_rate") else None,   ## 당시 환율
        "usd_pledged": round(float(p.get("usd_pledged") or 0), 2),  ## 당시 환율 적용값
        "usd_rate_current": round(float(p["fx_rate"]), 4) if p.get("fx_rate") else None, ## 수집 시점 환율
        "usd_pledged_current": round(float(p.get("pledged") or 0) * float(p["fx_rate"]), 2) if p.get("fx_rate") else None,  ## 수집 시점 환율 적용값 (페이지 표시)
        "percent_funded": g.get("percentFunded", p.get("percent_funded")),
        "has_photos": img_count > 0,
        "photo_count": img_count,
        "has_video": bool(g.get("video")) or video_count > 0,
        "story_video_count": video_count,
        "has_main_video": bool(g.get("video") or p.get("video")),
        "start_date": fmt_date(launched),
        "end_date": fmt_date(deadline),
        "duration_days": round((deadline - launched) / 86400, 1) if launched and deadline else None,
        "creator_name": (p.get("creator") or {}).get("name"),
        "creator_badges": ", ".join(badges),
        "creator_created_count": (creator_g.get("launchedProjects") or {}).get("totalCount"),
        "creator_backed_count": backed,
        "is_project_we_love": bool(g.get("isProjectWeLove", p.get("staff_pick"))),
        "updates_count": (g.get("posts") or {}).get("totalCount"),
        "comments_count": g.get("commentsCount"),
    }


def unique_path(path):
    """이미 같은 이름의 파일이 있으면 ' (1)', ' (2)'를 붙여 덮어쓰지 않는 경로를 돌려준다."""
    if not path.exists():
        return path
    n = 1
    while True:
        candidate = path.with_name(f"{path.stem} ({n}){path.suffix}")
        if not candidate.exists():
            return candidate
        n += 1


def run_export(fmt: str = "xlsx") -> None:
    config.OUTPUT_DIR.mkdir(exist_ok=True)
    with open(config.ENRICHED_FILE, encoding="utf-8") as f:
        projects = [json.loads(line) for line in f]
    date_str = datetime.now().strftime("%Y-%m-%d")

    frames: dict[str, pd.DataFrame] = {}
    for key, cat in config.CATEGORIES.items():
        rows = [build_row(p) for p in projects if p.get("_category_key") == key]
        if rows:
            frames[cat["name"]] = pd.DataFrame(rows, columns=COLUMNS)

    if fmt == "csv":
        for name, df in frames.items():
            path = unique_path(config.OUTPUT_DIR / f"kickstarter_{name.lower().replace(' ', '_')}_{date_str}.csv")
            df.to_csv(path, index=False, encoding="utf-8-sig")
            logger.info("[export] %s: %d행 → %s", name, len(df), path)
        return

    top_align = Alignment(vertical="top")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="4472C4", end_color="4472C4")
    excel_path = unique_path(config.OUTPUT_DIR / f"{config.EXCEL_BASENAME}_{date_str}.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        for name, df in frames.items():
            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.sheets[name]
            ws.freeze_panes = "A2"
            # story_text는 고정 너비, 나머지는 내용 길이에 맞춰 자동 조절
            for idx, col_name in enumerate(df.columns, start=1):
                letter = get_column_letter(idx)
                if col_name == "story_text":
                    width = 60
                else:
                    max_len = max(len(str(v)) for v in [col_name, *df[col_name].fillna("")])
                    width = min(max_len + 2, 100)
                ws.column_dimensions[letter].width = width
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = top_align
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            logger.info("[export] 시트 %s: %d행", name, len(df))
    logger.info("[export] 완료 → %s", excel_path)
